"""
sources/cbsl.py

Collector for Central Bank of Sri Lanka (CBSL) benchmark rate indicators.

Source: CBSL's official "4.04 Interest Rates - Monthly" statistical table,
linked from https://www.cbsl.gov.lk/en/statistics/statistical-tables/monetary-sector
as a downloadable .xlsx spreadsheet. This is the authoritative published
series for the Average Weighted Deposit Rate (AWDR), Average Weighted Fixed
Deposit Rate (AWFDR), and the policy rate, and is far more reliable to parse
than the bank's PDF monetary policy reviews, which present the same figures
as prose rather than a clean table.

The spreadsheet's download link is re-discovered on every run rather than
hardcoded: CBSL re-publishes this file monthly under a new, date-stamped
filename (e.g. "table4.04_20260609.xlsx"), so a fixed URL would silently
go stale.

Layout notes (confirmed by downloading and inspecting the live file):
  - Column index 1 (B): year. Only populated on the first row of each year;
    blank for subsequent months, so the value must be carried forward
    while scanning rows, the same pattern used for rowspanned HTML tables
    elsewhere in this project.
  - Column index 2 (C): month name.
  - Column index 3 (D): Overnight Policy Rate (OPR) — CBSL's current single
    policy rate, populated from late 2023 onward. Column index 4 (E),
    Standing Deposit Facility Rate (SDFR), is used as a fallback for older
    rows where OPR is blank, since SDFR was the operative signalling rate
    before OPR was introduced.
  - Column index 15 (P): AWDR. Column index 16 (Q): AWFDR.

No legal maximum deposit rate ("deposit_cap") is collected: CBSL does not
currently publish one as a standard recurring series — the cap imposed
during 2022/2023 was an emergency directive, not an ongoing published
statistic — so there is nothing reliable to scrape for that indicator yet.
"""

import io
import re
from datetime import datetime, timezone

import openpyxl
from bs4 import BeautifulSoup

from lib.http import fetch
from lib.robots import is_allowed
from lib.db import insert_cbsl, existing_cbsl_periods

STATISTICS_PAGE_URL = "https://www.cbsl.gov.lk/en/statistics/statistical-tables/monetary-sector"

_YEAR_COL, _MONTH_COL, _OPR_COL, _SDFR_COL = 1, 2, 3, 4
_AWDR_COL, _AWFDR_COL = 15, 16

# CBSL's month column is a full month name ("April"), not an abbreviation —
# %B in strptime handles that directly.
_PERIOD_FORMAT = "%B %Y"


def _find_monthly_interest_rates_url() -> str | None:
    """
    Fetch the CBSL monetary-sector statistics page and return the current
    download URL for the "Interest Rates - Monthly" spreadsheet, by matching
    on the link's visible text rather than a fixed, date-stamped filename.
    """
    response = fetch(STATISTICS_PAGE_URL)
    soup = BeautifulSoup(response.text, "html.parser")

    for link in soup.find_all("a", href=True):
        text = link.get_text(strip=True)
        if re.match(r"Interest Rates\s*-\s*Monthly", text, re.IGNORECASE):
            href = link["href"]
            return href if href.startswith("http") else f"https://www.cbsl.gov.lk{href}"
    return None


def _all_rows(xlsx_bytes: bytes) -> list[tuple[str, float | None, float, float]]:
    """
    Parse the spreadsheet and return every (period, policy_rate, awdr, awfdr)
    data row found, oldest first — the full multi-year series CBSL publishes
    in this one file, not just its most recent month.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    current_year = None
    rows = []

    for row in sheet.iter_rows(values_only=True):
        if row[_MONTH_COL] is None:
            continue
        if row[_YEAR_COL] is not None:
            current_year = row[_YEAR_COL]

        awdr  = row[_AWDR_COL]
        awfdr = row[_AWFDR_COL]
        if not isinstance(awdr, (int, float)) or not isinstance(awfdr, (int, float)):
            continue

        policy_rate = row[_OPR_COL] if isinstance(row[_OPR_COL], (int, float)) else row[_SDFR_COL]
        period = f"{row[_MONTH_COL]} {current_year}"
        rows.append((period, policy_rate, float(awdr), float(awfdr)))

    return rows


def _latest_row(xlsx_bytes: bytes) -> tuple[str, float | None, float, float] | None:
    """Return the most recent (period, policy_rate, awdr, awfdr), or None if the sheet has no data row."""
    rows = _all_rows(xlsx_bytes)
    return rows[-1] if rows else None


def _period_to_date(period: str) -> datetime:
    """
    Parse a CBSL period string (e.g. "April 2024") into a UTC datetime
    anchored to the first of that month, used as the synthetic scraped_at
    for backfilled historical rows — they need *some* date so
    getLatestBenchmarks()'s "most recent scraped_at wins" ordering still
    sorts them chronologically against each other and behind whatever a
    regular collect() run stores for the current month.
    """
    return datetime.strptime(period, _PERIOD_FORMAT).replace(tzinfo=timezone.utc)


def collect() -> int:
    """
    Fetch CBSL's monthly interest rates spreadsheet and store the latest
    AWDR, AWFDR, and policy rate values into cbsl_benchmarks.

    Returns:
        The number of indicator rows inserted (0 if the page is disallowed
        by robots.txt, the download link can't be found, or no data row is
        present in the spreadsheet).
    """
    if not is_allowed(STATISTICS_PAGE_URL):
        return 0

    xlsx_url = _find_monthly_interest_rates_url()
    if xlsx_url is None or not is_allowed(xlsx_url):
        return 0

    response = fetch(xlsx_url)
    latest = _latest_row(response.content)
    if latest is None:
        return 0

    period, policy_rate, awdr, awfdr = latest
    scraped_at = datetime.now(timezone.utc)

    insert_cbsl("awdr", awdr, period, xlsx_url, scraped_at)
    insert_cbsl("awfdr", awfdr, period, xlsx_url, scraped_at)
    count = 2

    if policy_rate is not None:
        insert_cbsl("policy_rate", float(policy_rate), period, xlsx_url, scraped_at)
        count += 1

    return count


def backfill() -> int:
    """
    One-time historical load: store every month CBSL's spreadsheet has data
    for (years deep), not just the latest. Run manually via
    backfill_cbsl_history.py — collect() (the regular 6-hourly run) only
    ever stores the latest month, so this is what gives the AWFDR/AWDR
    charts real multi-year depth instead of however many days the regular
    scrape has happened to run for.

    Idempotent: skips any (indicator, period) pair already in
    cbsl_benchmarks, so re-running this (or running it after collect() has
    already stored the current month) never creates duplicate rows.

    Returns:
        The number of indicator rows actually inserted.
    """
    if not is_allowed(STATISTICS_PAGE_URL):
        return 0

    xlsx_url = _find_monthly_interest_rates_url()
    if xlsx_url is None or not is_allowed(xlsx_url):
        return 0

    response = fetch(xlsx_url)
    rows = _all_rows(response.content)

    already_awdr  = existing_cbsl_periods("awdr")
    already_awfdr = existing_cbsl_periods("awfdr")
    already_policy = existing_cbsl_periods("policy_rate")

    count = 0
    for period, policy_rate, awdr, awfdr in rows:
        scraped_at = _period_to_date(period)

        if period not in already_awdr:
            insert_cbsl("awdr", awdr, period, xlsx_url, scraped_at)
            count += 1
        if period not in already_awfdr:
            insert_cbsl("awfdr", awfdr, period, xlsx_url, scraped_at)
            count += 1
        if policy_rate is not None and period not in already_policy:
            insert_cbsl("policy_rate", float(policy_rate), period, xlsx_url, scraped_at)
            count += 1

    return count
